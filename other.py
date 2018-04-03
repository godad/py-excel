import openpyxl
from collections import OrderedDict
from openpyxl.styles import Alignment
from openpyxl.styles.fills import PatternFill


__author__ = "Guguangyu"
__date__ = "2017-5-31"

# Y
wby = openpyxl.load_workbook('autoio.xlsx')
wsy = wby.worksheets[0]
rowYMax = wsy.max_row + 1

# F
wbf = openpyxl.load_workbook('hardware.xlsx')
wsf = wbf.worksheets[0]
rowFMax = wsf.max_row + 1

dictA = OrderedDict()
dictB = OrderedDict()
dictAB = OrderedDict()
dictBA = OrderedDict()

# 新建excel
wb = openpyxl.load_workbook('newfile.xlsx')
sheet = wb.active
sheet.title = "materials"
sheet['A1'] = 'Location'
sheet['B1'] = 'ProductID'
sheet['C1'] = 'AutoIoID'
sheet['D1'] = 'Value'
sheet['E1'] = 'Description'
sheet['F1'] = 'Quantity'
sheet['G1'] = 'Brand'

align = Alignment(horizontal="center", vertical="center")

# 易工模板
for rowsY in range(2, rowYMax):
    autoIoy = wsy['B' + str(rowsY)].value
    factory = wsy['E' + str(rowsY)].value
    brandy = wsy['F' + str(rowsY)].value
    dictA[factory] = autoIoy
    dictAB[factory] = brandy
# print(dictAB)

# 生成友衷编码
for rowsF in range(2, rowFMax):
    autoIof = wsf['E' + str(rowsF)].value
    dictB[autoIof] = " "
    dictBA[autoIof] = " "
for keyB in dictB.keys():
    for keyA in dictA.keys():
        if keyB == keyA:
            dictB[keyB] = dictA[keyA]

# 生成厂商目录
for keyBA in dictBA.keys():
    for keyAB in dictAB.keys():
        if keyBA == keyAB:
            dictBA[keyBA] = dictAB[keyAB]

fill = PatternFill(patternType='mediumGray')
fillSolid = PatternFill(patternType='lightUp')

# 产生新的列表
for rowN in range(2, rowFMax):
    sheet['A' + str(rowN)].value = wsf['C' + str(rowN)].value
    sheet['B' + str(rowN)].value = wsf['E' + str(rowN)].value
    sheet['C' + str(rowN)].value = dictB[wsf['E' + str(rowN)].value]
    sheet['E' + str(rowN)].value = wsf['H' + str(rowN)].value

    # 计算物料总量
    number = str(sheet['A' + str(rowN)].value)
    sheet['F' + str(rowN)].value = number.count(',') + 1
    sheet['G' + str(rowN)].value = dictBA[wsf['E' + str(rowN)].value]
    sheet['D' + str(rowN)].value = wsf['D' + str(rowN)].value

    # 检测是否存在友衷物料编码
    colDvalue = sheet.row_dimensions[rowN]
    colDvalue.alignment = align
    if sheet['C' + str(rowN)].value == " ":
        colDvalue.fill = fill

# 硬件提交表单检测
wbfValue = []
for i in wsf["E"]:
    wbfValue.append(i.value)

productL = list(wbfValue)
productS = list(set(wbfValue))
productS.sort(key=productL.index)

pos = []
for i in productS:
    num = productL.count(i)
    if num > 1:
        def myfind(x, y):
            return [a for a in range(len(y)) if y[a] == x]
        position = myfind(i, productL)
        pos.append(position)
        length = len(pos)

        # if len(position) > 1 and len(position) < 3:
listNew = []
for i in range(0, length):
    listA = [str(pos[i][0] + 1), sheet['D' + str(pos[i][0] + 1)].value]
    listB = [str(pos[i][1] + 1), sheet['D' + str(pos[i][1] + 1)].value]
    listNew.append(listA)
    listNew.append(listB)
# print(listNew)

output = len(listNew)
if output == 0:
    file = open('output', 'w')
    file.write(str(file))
    file.close()

# 重复编码写入文件中
file = open('result.txt', 'w')
file.write(str(listNew))
file.close()


wb.save('newfile.xlsx')